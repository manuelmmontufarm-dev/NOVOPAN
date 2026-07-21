"""Genera el Excel de solicitud de tags HMI para el área de Sistemas.

Se corre desde `simulador-final/`:  python3 scripts/gen-excel-sistemas.py

El .xlsx es un ENTREGABLE versionado, pero se regenera desde aquí — no se
edita a mano. Cuando Sistemas responda los pendientes de la hoja
"Por confirmar", se actualizan las filas de `rows`/`pend` y se vuelve a
correr, igual que scripts/gen-tags.mjs con TAGS.md.

⚠️ Tras correr esto hay que RECALCULAR el libro (LibreOffice o abrir/guardar
en Excel): openpyxl escribe las fórmulas del resumen sin valor cacheado, y
sin recalcular esas cinco celdas se ven vacías para quien lo reciba.

Alcance actual: SOLO Línea 1 Sección 2 (silos 5/6 → sensores). El tramo de
preparación está en MAPEO-WINCC.md y se pedirá en una segunda tanda.
"""
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F = "Arial"
VERDE = "1B5E20"      # encabezados
GRIS = "F2F2F2"
AMARILLO = "FFF3CD"   # celdas que Sistemas debe llenar / confirmar
ROJO_TXT = "B00020"

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ─────────────────────────── Hoja 1 · Instrucciones ───────────────────────────
ws = wb.active
ws.title = "Instrucciones"
ws.sheet_view.showGridLines = False

def put(cell, val, *, size=11, bold=False, color="000000", wrap=False, fill=None):
    c = ws[cell]
    c.value = val
    c.font = Font(name=F, size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 88

put("B2", "NOVOPAN · Línea 1 — Tags del HMI para el simulador de trazabilidad", size=15, bold=True, color=VERDE)
put("B3", "Solicitud al área de Sistemas · 21-jul-2026", size=10, color="666666")

put("B5", "Qué se necesita", size=12, bold=True, color=VERDE)
put("B6", "En una frase", bold=True)
put("C6", "Que un job escriba tres archivos de texto con los valores de los tags que ya existen en "
          "los HMI. No hay que programar ningún servicio web, API ni endpoint.", wrap=True)

put("B8", "Dónde se escriben", bold=True)
put("C8", "En la carpeta datos/ del sitio del simulador. Un archivo por servidor WinCC:", wrap=True)
put("C9", "datos/hmi-formacion.csv      ←  servidor de Formación (HMI-METSO, 700 tags)")
put("C10", "datos/hmi-encolado.csv       ←  servidor de Encolado / cocina de cola (420 tags)")
put("C11", "datos/hmi-preparacion.csv    ←  servidor de Preparación / secado (586 tags)")
for r in (9, 10, 11):
    ws[f"C{r}"].font = Font(name=F, size=10)

put("B13", "Por qué tres y no uno", bold=True)
put("C13", "Dos servidores publican un tag con el MISMO nombre y distinto significado "
           "(H_PressSpeed_PV es \"Press speed\" en Formación y \"Press Conveyor Speed\" en Encolado). "
           "Si van en un solo archivo, uno pisa al otro. Separados, el simulador sabe cuál es cuál.", wrap=True)
ws.row_dimensions[13].height = 42

put("B15", "Cada cuánto", bold=True)
put("C15", "El simulador relee los archivos cada 2 segundos. Si el job puede escribir más lento "
           "(5 s, 10 s, 1 min) también sirve — nada se rompe, solo baja la resolución.", wrap=True)
ws.row_dimensions[15].height = 30

put("B17", "Formato del archivo", size=12, bold=True, color=VERDE)
put("B18", "Una línea por tag", bold=True)
put("C18", "NOMBRE_DEL_TAG: valor;      ← dos puntos separan, punto y coma termina", wrap=True)
put("B19", "Nombres", bold=True)
put("C19", "EXACTAMENTE como aparecen en WinCC. No hay que renombrar nada — el simulador traduce. "
           "Se respetan mayúsculas, guiones y hasta los errores de tipeo de fábrica.", wrap=True)
ws.row_dimensions[19].height = 30
put("B20", "Decimal", bold=True)
put("C20", "Punto o coma, los dos sirven (14.5 y 14,5 son iguales).")
put("B21", "Comentarios", bold=True)
put("C21", "Líneas que empiezan con # se ignoran.")
put("B22", "Tag pendiente", bold=True)
put("C22", "Dejarlo con el valor vacío (NOMBRE_DEL_TAG:;). El simulador lo marca como pendiente "
           "y no rompe nada.", wrap=True)
put("B23", "Tag que no conozca", bold=True)
put("C23", "Se ignora con un aviso. Mandar de más NO rompe el archivo — si tienen dudas, "
           "es preferible incluirlo.", wrap=True)

put("B25", "Ejemplo real de datos/hmi-formacion.csv", size=12, bold=True, color=VERDE)
ej = [
    "# NOVOPAN L1 · Formación · generado 2026-07-21 10:15:03",
    "H_PressSpeed_PV: 14.5;",
    "H_Act_MatWeight_SP: 11.5;",
    "H_CL_Total_Flakes: 118;",
    "H_Act_SL1_SP: 47.1;",
    "H_FL_FormConv_Speed_PV:;      # todavía no lo estamos mandando",
]
for i, line in enumerate(ej):
    c = ws.cell(row=26 + i, column=3, value=line)
    c.font = Font(name="Courier New", size=10, color="000000" if i else "666666")
    c.fill = PatternFill("solid", fgColor=GRIS)

put("B34", "Resumen de lo pedido", size=12, bold=True, color=VERDE)
res = [
    ("Total de tags solicitados", '=COUNTA(Tags!C2:C100)'),
    ("Prioridad 1 · Crítico", '=COUNTIF(Tags!G2:G100,"1 · Crítico")'),
    ("Prioridad 2 · Deseado", '=COUNTIF(Tags!G2:G100,"2 · Deseado")'),
    ("Prioridad 3 · Opcional", '=COUNTIF(Tags!G2:G100,"3 · Opcional")'),
    ("Puntos por confirmar", '=COUNTA(\'Por confirmar\'!B2:B40)'),
]
for i, (label, formula) in enumerate(res):
    r = 35 + i
    put(f"B{r}", label, size=10, bold=(i == 0))
    c = ws.cell(row=r, column=3, value=formula)
    c.font = Font(name=F, size=10, bold=(i == 0))
    c.alignment = Alignment(horizontal="left")

put("B41", "Si algo no se entiende, escribir a Manuel Montúfar (pasantía de ingeniería) "
           "antes de improvisar un nombre de tag.", size=10, color="666666", wrap=True)

# ─────────────────────────────── Hoja 2 · Tags ────────────────────────────────
ws = wb.create_sheet("Tags")
cols = [
    ("Servidor", 15),
    ("Archivo CSV destino", 26),
    ("Tag en WinCC (nombre exacto)", 32),
    ("Access Name", 13),
    ("Qué es", 46),
    ("Unidad", 11),
    ("Prioridad", 14),
    ("Nota", 52),
]
for i, (name, width) in enumerate(cols, start=1):
    c = ws.cell(row=1, column=i, value=name)
    c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VERDE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = box
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 30

A, B, C = "Formación", "Encolado", "Preparación"
FA, FB, FC = "datos/hmi-formacion.csv", "datos/hmi-encolado.csv", "datos/hmi-preparacion.csv"
P1, P2, P3 = "1 · Crítico", "2 · Deseado", "3 · Opcional"

rows = [
    # ── Servidor A · Formación ──
    (A, FA, "H_PressSpeed_PV", "Forming", "Velocidad de prensa", "m/min", P1, "Confirmado. Es el máster de toda la línea."),
    (A, FA, "H_Act_MatWeight_SP", "Forming", "Peso de manta después de formación", "kg/m²", P1, "Confirmado. OJO: NO usar H_Act_MatWeight_Real, ese está en kg (báscula), no en kg/m²."),
    (A, FA, "H_CL_Total_Flakes", "Forming", "Flujo total de capa gruesa (core · CL)", "kg/min", P1, "Confirmado."),
    (A, FA, "H_Act_SL1_SP", "Forming", "% de receta de la capa fina inferior (SL1)", "%", P1, "Confirmado."),
    (A, FA, "H_FL_FormConv_Speed_PV", "Forming", "Velocidad de la banda de formación", "m/min", P2, "Hoy el simulador la asume igual a la de prensa. Con este tag deja de suponer."),
    (A, FA, "H_FL_FormConv_Speed_Mode", "Forming", "Banda sigue a la prensa (1) o va en manual (0)", "0/1", P2, ""),
    (A, FA, "H_PP_Upper_Speed_PV", "Forming", "Velocidad de pre-prensa, accionamiento superior", "m/min", P2, ""),
    (A, FA, "H_PP_Lower_Speed_PV", "Forming", "Velocidad de pre-prensa, accionamiento inferior", "m/min", P2, ""),
    (A, FA, "H_SL1_Filling_PV", "Forming", "Llenado de la tolva del esparcidor SL1", "%", P2, ""),
    (A, FA, "H_CC_Filling_PV", "Forming", "Llenado de la tolva del esparcidor de core", "%", P2, ""),
    (A, FA, "H_SL2_Filling_PV", "Forming", "Llenado de la tolva del esparcidor SL2", "%", P2, ""),
    (A, FA, "H_Empty_ON", "Forming", "Vaciado de esparcidores en curso", "0/1", P2, "Sirve para calibrar cuántos kg cabe en cada tolva sin tener que buscarlo en planos."),
    (A, FA, "H_MaterialFlow_TonsPerHour", "Forming", "Flujo total de material de la línea", "t/h", P3, "Validación cruzada contra los flujos por capa."),
    (A, FA, "H_Moisure_PV", "Forming", "Humedad (instrumento PEW758)", "%", P3, ""),
    (A, FA, "H_TrimSaw_Width", "Forming", "Ancho de refilado", "mm", P3, ""),
    (A, FA, "IMAN_AUTO", "Forming", "Imán en automático", "0/1", P3, ""),
    (A, FA, "P_Act_LineSpeed_SP", "Press", "Velocidad de línea", "m/min", P3, ""),
    (A, FA, "Q_Act_PanelLength", "POF", "Largo del tablero terminado", "mm", P3, ""),
    (A, FA, "Q_Act_PanelWidth", "POF", "Ancho del tablero terminado", "mm", P3, ""),
    (A, FA, "Q_Act_FinalThickness", "POF", "Espesor final del tablero", "mm", P3, ""),
    (A, FA, "Q_Act_PanelAmount", "POF", "Conteo de tableros de la orden", "u", P3, ""),

    # ── Servidor B · Encolado ──
    (B, FB, "F_SL_FlakeFlow_PV", "Gluing", "Flujo total de capa fina (SL)", "kg/min ⚠", P1, "CONFIRMAR UNIDAD: el comentario del HMI no la declara. Asumimos kg/min por simetría."),
    (B, FB, "F_CL_DosBin_Weight", "Gluing", "Peso en la tolva dosificadora gruesa", "kg", P1, "Confirmado."),
    (B, FB, "F_SL_DosBin_Weight", "Gluing", "Peso en la tolva dosificadora fina", "kg", P1, "Confirmado."),
    (B, FB, "F_CL_Filling_PV", "Gluing", "Nivel de la tolva dosificadora gruesa", "%", P2, ""),
    (B, FB, "F_SL_Filling_PV", "Gluing", "Nivel de la tolva dosificadora fina", "%", P2, ""),
    (B, FB, "OK_BOMBAS_GLUING_CL", "Gluing", "Bombas de encolado grueso", "0/1 ⚠", P2, "CONFIRMAR POLARIDAD: el nombre dice OK pero el comentario dice \"FALLA BOMBAS ENCOLADO CL\"."),
    (B, FB, "OK_BOMBAS_GLUING_SL", "Gluing", "Bombas de encolado fino", "0/1 ⚠", P2, "CONFIRMAR POLARIDAD (igual que el anterior)."),
    (B, FB, "F_CL_FlakeDens_PV", "Gluing", "Densidad de flakes en la dosificadora gruesa", "kg/m³", P1, "Se usa como densidad del silo 5 (core): el bin se llena directo del silo con el mismo material."),
    (B, FB, "F_SL_FlakeDens_PV", "Gluing", "Densidad de flakes en la dosificadora fina", "kg/m³", P1, "Se usa como densidad del silo 6 (capa fina)."),
    (B, FB, "F_CL_FlakeFlow_PV", "Gluing", "Flujo de capa gruesa", "kg/min", P3, "Redundante con H_CL_Total_Flakes. Sirve para contrastar."),
    (B, FB, "F_CL_DosBin_Speed", "Gluing", "Velocidad de la banda dosificadora gruesa", "—", P3, ""),
    (B, FB, "F_SL_DosBin_Speed", "Gluing", "Velocidad de la banda dosificadora fina", "—", P3, ""),
    (B, FB, "— NO INCLUIR —", "Gluing", "H_PressSpeed_PV y H_SL_FlakeDens_SP", "—", "", "Estos dos existen también en Formación con otro significado. Mandarlos desde acá causa conflicto: NO ponerlos en este archivo."),

    # ── Servidor C · Preparación ──
    (C, FC, "066_C_Dry_Materia… (¿?)", "SCRN", "Silo 5 (capa interna · core) — nivel", "%", P1, "NOMBRE INCOMPLETO: clic en el tag y leer la barra de estado abajo del diálogo. Comment: \"Nivel silo capa interna %\". Alimenta Sección 2."),
    (C, FC, "066_C_Dry_Materia… (¿?)", "SCRN", "Silo 5 (capa interna · core) — descarga", "kg/h", P1, "NOMBRE INCOMPLETO (ídem). Comment: \"Descarga desde silo capa interna kg/h\". OJO: el simulador lo usa en kg/min — mandar el valor crudo en kg/h, nosotros convertimos."),
    (C, FC, "066_C_Dry_Materia… (¿?)", "SCRN", "Silo 6 (capa externa · fina) — nivel", "%", P1, "NOMBRE INCOMPLETO (ídem). Comment: \"Nivel silo capa externa %\". Alimenta Sección 2."),
    (C, FC, "066_C_Dry_Materia… (¿?)", "SCRN", "Silo 6 (capa externa · fina) — descarga", "kg/h", P1, "NOMBRE INCOMPLETO (ídem). Comment: \"Descarga desde silo capa externa kg/h\". Mandar crudo en kg/h."),
    (C, FC, "Prensa_metal_detector", "Prensa", "Metal detectado en el material (detector a 37.69 m)", "0/1", P2, "Está en el servidor de Preparación aunque el equipo es de Sección 2."),
]
# Alcance deliberado: SOLO Sección 2 (silos 5/6 → sensores). El tramo de
# preparación (silos húmedos, bunker, secadero) queda documentado en
# MAPEO-WINCC.md y se pedirá en una segunda tanda — pedir todo junto
# retrasaría los datos que el simulador puede usar HOY.

for r, row in enumerate(rows, start=2):
    for i, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.font = Font(name=F, size=10)
        c.alignment = Alignment(wrap_text=(i in (5, 8)), vertical="top")
        c.border = box
    if row[0] == A:
        tint = "FFFFFF"
    elif row[0] == B:
        tint = "F7F9F7"
    else:
        tint = "EEF3EE"
    for i in range(1, 9):
        ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=tint)
    ws.cell(row=r, column=3).font = Font(name="Courier New", size=10, bold=True)
    if row[6] == P1:
        ws.cell(row=r, column=7).font = Font(name=F, size=10, bold=True, color=ROJO_TXT)
    if "⚠" in str(row[5]) or "CONFIRMAR" in str(row[7]) or "NO INCLUIR" in str(row[2]):
        ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor=AMARILLO)
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=AMARILLO)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

# ───────────────────────── Hoja 3 · Por confirmar ─────────────────────────────
ws = wb.create_sheet("Por confirmar")
cols3 = [("Tema", 22), ("Qué necesitamos", 62), ("A quién", 20), ("Por qué importa", 60)]
for i, (name, width) in enumerate(cols3, start=1):
    c = ws.cell(row=1, column=i, value=name)
    c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VERDE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = box
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 26

# Solo lo que bloquea Línea 1 · Sección 2. Lo del tramo de preparación
# (nombres de descargas de silos húmedos, Wet_Mat_CALC, unidades de Cap/Level,
# correspondencia aserrín↔Silo1, ambigüedad 041.11) queda documentado en
# MAPEO-WINCC.md y se pedirá cuando se cablee ese tramo.
pend = [
    ("Nombre incompleto", "Los 4 tags que empiezan con 066_C_Dry_Materia… (Access SCRN). Sus comentarios son "
     "\"Nivel silo capa interna %\", \"Descarga desde silo capa interna kg/h\", \"Nivel silo capa externa %\" "
     "y \"Descarga desde silo capa externa kg/h\". Necesitamos el nombre completo de los cuatro "
     "(clic en el tag → el nombre completo sale en la barra de estado abajo del diálogo).",
     "Sistemas", "Son los silos 5 (core) y 6 (capa fina) — la entrada de Sección 2. Es lo más importante de esta lista."),
    ("Unidad", "¿En qué unidad está F_SL_FlakeFlow_PV? Su comentario dice solo \"SL flake flow\".",
     "Sistemas", "Lo estamos usando como kg/min. Si es kg/h, el modelo se equivoca por 60×."),
    ("Polaridad", "OK_BOMBAS_GLUING_CL y _SL: ¿1 significa que están OK, o que hay falla? El nombre dice una cosa "
     "y el comentario del HMI dice la contraria.",
     "Sistemas", "Al revés, el simulador marcaría la línea parada justo cuando está corriendo."),
]

for r, row in enumerate(pend, start=2):
    for i, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.font = Font(name=F, size=10, bold=(i == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = box
        c.fill = PatternFill("solid", fgColor=AMARILLO if i == 2 else "FFFFFF")
    ws.row_dimensions[r].height = 52

ws.freeze_panes = "A2"

out = pathlib.Path(__file__).resolve().parent.parent / "NOVOPAN_L1_Tags_HMI_para_Sistemas.xlsx"
wb.save(out)
print("guardado:", out)
print("tags:", len(rows), "· pendientes:", len(pend))
